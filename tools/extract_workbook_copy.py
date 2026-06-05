from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_WORKBOOK = Path(
    r"B:\PERSONAL\HYDROPONICS\outputs\hydroponics_financial_statement_template\hydroponics_financial_statement_owner_model_v3_pack_sales.xlsx"
)
OUTPUT_JS = Path(__file__).resolve().parents[1] / "src" / "workbook-copy.js"
APP_VERSION = "1.0.0"


def clean(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def month(value: Any) -> str:
    value = clean(value)
    if not value:
        return ""
    text = str(value)
    return text[:7] if len(text) >= 7 else text


def number(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(clean(value))


def is_present(value: Any) -> bool:
    return value not in (None, "", 0, "0")


def get_rows(workbook, sheet_name: str) -> list[list[Any]]:
    ws = workbook[sheet_name]
    return [[clean(cell.value) for cell in row] for row in ws.iter_rows()]


def cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def setup_assumptions(rows: list[list[Any]]) -> dict[str, Any]:
    lookup = {text(cell(row, 1)).lower(): clean(cell(row, 2)) for row in rows}
    return {
        "fiscalYear": int(number(lookup.get("fiscal year"), 2026)),
        "businessName": text(lookup.get("business name")) or "Hydroponics Lettuce Business",
        "owner": text(lookup.get("owner")) or "Yhan",
        "location": text(lookup.get("location")) or "Bohol, Philippines",
        "currency": text(lookup.get("currency")) or "PHP",
        "defaultSellingUnit": text(lookup.get("default selling unit")) or "pack",
        "targetGrossMargin": number(lookup.get("target gross margin"), 0.5),
        "incomeTaxRate": number(lookup.get("income tax rate / provision"), 0),
        "beginningCash": number(lookup.get("beginning cash"), 0),
        "defaultPricePack": number(lookup.get("default selling price per pack"), 35),
        "defaultLaborRate": number(lookup.get("default hourly labor rate"), 62.5),
        "defaultElectricityRate": number(lookup.get("default kwh rate"), 11),
        "defaultLettucesPerPack": number(lookup.get("lettuces per pack"), 3),
    }


def purchase_rows(rows: list[list[Any]], ledger: str) -> list[dict[str, Any]]:
    records = []
    for idx, row in enumerate(rows[4:], start=5):
        date_value = clean(cell(row, 0))
        item = text(cell(row, 1))
        if not date_value or not item:
            continue
        if ledger == "nutrient":
            category = "Nutrient"
            form = text(cell(row, 2))
            unit = form
            qty = number(cell(row, 3))
            total_cost = number(cell(row, 4))
            supplier = text(cell(row, 6))
            paid = text(cell(row, 8)) or "No"
            remarks = text(cell(row, 9))
        else:
            category = text(cell(row, 2))
            form = text(cell(row, 3))
            unit = form
            qty = number(cell(row, 4))
            total_cost = number(cell(row, 5))
            supplier = text(cell(row, 7))
            paid = text(cell(row, 8)) or "No"
            remarks = text(cell(row, 9))
        if qty == 0 and total_cost == 0:
            continue
        records.append(
            {
                "id": f"{ledger}-purchase-excel-r{idx}",
                "ledger": ledger,
                "date": date_value,
                "item": item,
                "category": category,
                "form": form,
                "unit": unit,
                "qty": qty,
                "totalCost": total_cost,
                "supplier": supplier,
                "paid": paid,
                "remarks": remarks,
            }
        )
    return records


def usage_rows(rows: list[list[Any]], ledger: str) -> list[dict[str, Any]]:
    records = []
    for idx, row in enumerate(rows[4:], start=5):
        date_value = clean(cell(row, 0))
        if ledger == "nutrient":
            item = text(cell(row, 5))
            if not date_value or not item or item == "0":
                continue
            qty = number(cell(row, 7))
            records.append(
                {
                    "id": f"{ledger}-usage-excel-r{idx}",
                    "ledger": ledger,
                    "date": date_value,
                    "batch": text(cell(row, 2)),
                    "tankRef": text(cell(row, 3)),
                    "eventType": text(cell(row, 4)),
                    "item": item,
                    "category": "Nutrient",
                    "unitUsed": text(cell(row, 6)),
                    "qtyUsed": qty,
                    "purpose": text(cell(row, 10)),
                    "remarks": text(cell(row, 11)),
                }
            )
        else:
            item = text(cell(row, 3))
            if not date_value or not item or item == "0":
                continue
            qty = number(cell(row, 6))
            records.append(
                {
                    "id": f"{ledger}-usage-excel-r{idx}",
                    "ledger": ledger,
                    "date": date_value,
                    "batch": text(cell(row, 2)),
                    "tankRef": "",
                    "eventType": "",
                    "item": item,
                    "category": text(cell(row, 4)),
                    "unitUsed": text(cell(row, 5)),
                    "qtyUsed": qty,
                    "purpose": text(cell(row, 9)),
                    "remarks": text(cell(row, 10)),
                }
            )
    return records


def build_structured_state(values_wb) -> dict[str, Any]:
    setup = setup_assumptions(get_rows(values_wb, "Setup & Assumptions"))
    sales = []
    for idx, row in enumerate(get_rows(values_wb, "Sales Register")[4:], start=5):
        date_value = clean(cell(row, 0))
        crop = text(cell(row, 2))
        if not date_value or not crop:
            continue
        sales.append(
            {
                "id": f"sale-excel-r{idx}",
                "date": date_value,
                "crop": crop,
                "batch": text(cell(row, 3)),
                "packsSold": number(cell(row, 4)),
                "lettucesPerPack": number(cell(row, 5), setup["defaultLettucesPerPack"]),
                "pricePack": number(cell(row, 7), setup["defaultPricePack"]),
                "discountReturns": number(cell(row, 9)),
                "collectionStatus": text(cell(row, 11)) or "Collected",
                "remarks": text(cell(row, 12)),
            }
        )

    purchases = []
    purchases += purchase_rows(get_rows(values_wb, "Nutrient Purchases"), "nutrient")
    purchases += purchase_rows(get_rows(values_wb, "Seed & Supply Purchases"), "supply")
    purchases += purchase_rows(get_rows(values_wb, "Chemical Purchases"), "chemical")

    usages = []
    usages += usage_rows(get_rows(values_wb, "Nutrient Mixing Usage"), "nutrient")
    usages += usage_rows(get_rows(values_wb, "Seed & Supply Usage"), "supply")
    usages += usage_rows(get_rows(values_wb, "Chemical Usage"), "chemical")

    tanks = []
    for idx, row in enumerate(get_rows(values_wb, "Tank & Top-up Log")[4:], start=5):
        date_value = clean(cell(row, 0))
        tank_ref = text(cell(row, 2))
        if not date_value or not tank_ref:
            continue
        tanks.append(
            {
                "id": f"tank-excel-r{idx}",
                "date": date_value,
                "tankRef": tank_ref,
                "batch": text(cell(row, 3)),
                "tankId": text(cell(row, 4)),
                "eventType": text(cell(row, 5)),
                "startingVolumeL": number(cell(row, 6)),
                "waterAddedL": number(cell(row, 7)),
                "waterDischargedL": number(cell(row, 8)),
                "endingVolumeL": number(cell(row, 9)),
                "ec": number(cell(row, 10)),
                "ph": number(cell(row, 11)),
                "notes": text(cell(row, 14)),
            }
        )

    labor = []
    for idx, row in enumerate(get_rows(values_wb, "Labor")[4:], start=5):
        date_value = clean(cell(row, 0))
        worker = text(cell(row, 2))
        if not date_value or not worker:
            continue
        labor.append(
            {
                "id": f"labor-excel-r{idx}",
                "date": date_value,
                "workerRole": worker,
                "task": text(cell(row, 3)),
                "hoursWorked": number(cell(row, 4)),
                "overrideRateHour": number(cell(row, 5), ""),
                "production": text(cell(row, 8)) or "Production",
                "batch": text(cell(row, 9)),
                "remarks": text(cell(row, 10)),
            }
        )

    electricity = []
    for idx, row in enumerate(get_rows(values_wb, "Electricity")[4:], start=5):
        month_value = clean(cell(row, 5))
        device = text(cell(row, 6))
        if not month_value or not device:
            continue
        electricity.append(
            {
                "id": f"power-excel-r{idx}",
                "month": month(month_value),
                "device": device,
                "ratedWatts": number(cell(row, 7)),
                "qty": number(cell(row, 8), 1),
                "hoursDay": number(cell(row, 9)),
                "daysUsed": number(cell(row, 10)),
                "rateKwh": number(cell(row, 12), setup["defaultElectricityRate"]),
                "production": text(cell(row, 14)) or "Production",
                "batch": text(cell(row, 15)),
                "remarks": text(cell(row, 16)),
            }
        )

    assets = []
    for idx, row in enumerate(get_rows(values_wb, "Fixed Assets")[4:], start=5):
        date_value = clean(cell(row, 0))
        description = text(cell(row, 2))
        if not date_value or not description:
            continue
        assets.append(
            {
                "id": f"asset-excel-r{idx}",
                "acquisitionDate": date_value,
                "category": text(cell(row, 1)),
                "description": description,
                "qty": number(cell(row, 3), 1),
                "totalCost": number(cell(row, 4)),
                "usefulLifeMonths": number(cell(row, 5), 1),
                "salvageValue": number(cell(row, 6)),
                "depStartMonth": month(cell(row, 10)),
                "depEndMonth": month(cell(row, 11)),
                "status": text(cell(row, 12)),
                "remarks": text(cell(row, 13)),
            }
        )

    expenses = []
    for idx, row in enumerate(get_rows(values_wb, "Other Expenses")[4:], start=5):
        date_value = clean(cell(row, 0))
        description = text(cell(row, 3))
        if not date_value or not description:
            continue
        expenses.append(
            {
                "id": f"expense-excel-r{idx}",
                "date": date_value,
                "category": text(cell(row, 2)),
                "description": description,
                "amount": number(cell(row, 4)),
                "production": text(cell(row, 5)) or "Overhead",
                "paid": text(cell(row, 6)) or "No",
                "receiptRef": text(cell(row, 7)),
                "batch": text(cell(row, 8)),
                "remarks": text(cell(row, 9)),
            }
        )

    financing = []
    for idx, row in enumerate(get_rows(values_wb, "Financing & Equity")[4:], start=5):
        date_value = clean(cell(row, 0))
        type_value = text(cell(row, 2))
        if not date_value or not type_value:
            continue
        financing.append(
            {
                "id": f"finance-excel-r{idx}",
                "date": date_value,
                "type": type_value,
                "amount": number(cell(row, 3)),
                "counterparty": text(cell(row, 4)),
                "receiptRef": text(cell(row, 5)),
                "cashEffect": number(cell(row, 6)),
                "remarks": text(cell(row, 7)),
            }
        )

    cycles = []
    for idx, row in enumerate(get_rows(values_wb, "Crop Cycle Costing")[4:], start=5):
        batch = text(cell(row, 0))
        if not batch:
            continue
        cycles.append(
            {
                "id": f"cycle-excel-r{idx}",
                "batch": batch,
                "crop": text(cell(row, 1)),
                "startDate": clean(cell(row, 2)) or "",
                "harvestMonth": month(cell(row, 3)),
                "expectedPacks": number(cell(row, 4)),
                "status": text(cell(row, 21)),
                "remarks": "Copied from Excel workbook.",
            }
        )

    return {
        "appVersion": APP_VERSION,
        "updatedAt": "",
        "sourceWorkbook": {
            "fileName": SOURCE_WORKBOOK.name,
            "folder": str(SOURCE_WORKBOOK.parent),
            "lastModified": datetime.fromtimestamp(SOURCE_WORKBOOK.stat().st_mtime).isoformat(timespec="seconds"),
        },
        "assumptions": setup,
        "sales": sales,
        "purchases": purchases,
        "usages": usages,
        "tanks": tanks,
        "labor": labor,
        "electricity": electricity,
        "assets": assets,
        "expenses": expenses,
        "financing": financing,
        "cycles": cycles,
    }


def workbook_copy(formulas_wb) -> list[dict[str, Any]]:
    sheets = []
    for ws in formulas_wb.worksheets:
        rows = [[clean(cell.value) for cell in row] for row in ws.iter_rows()]
        sheets.append(
            {
                "name": ws.title,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "rows": rows,
            }
        )
    return sheets


def main() -> None:
    formulas_wb = load_workbook(SOURCE_WORKBOOK, data_only=False, read_only=True)
    values_wb = load_workbook(SOURCE_WORKBOOK, data_only=True, read_only=True)
    payload = {
        "source": {
            "fileName": SOURCE_WORKBOOK.name,
            "folder": str(SOURCE_WORKBOOK.parent),
            "lastModified": datetime.fromtimestamp(SOURCE_WORKBOOK.stat().st_mtime).isoformat(timespec="seconds"),
            "extractedAt": datetime.now().isoformat(timespec="seconds"),
        },
        "sheets": workbook_copy(formulas_wb),
        "state": build_structured_state(values_wb),
    }
    formulas_wb.close()
    values_wb.close()

    OUTPUT_JS.write_text(
        "export const WORKBOOK_SOURCE = "
        + json.dumps(payload["source"], ensure_ascii=True, separators=(",", ":"))
        + ";\nexport const WORKBOOK_COPY = "
        + json.dumps(payload["sheets"], ensure_ascii=True, separators=(",", ":"))
        + ";\nexport const EXCEL_IMPORTED_STATE = "
        + json.dumps(payload["state"], ensure_ascii=True, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    counts = {key: len(payload["state"].get(key, [])) for key in ["sales", "purchases", "usages", "tanks", "labor", "electricity", "assets", "expenses", "financing", "cycles"]}
    print(json.dumps({"output": str(OUTPUT_JS), "sheetCount": len(payload["sheets"]), "counts": counts}, indent=2))


if __name__ == "__main__":
    main()
